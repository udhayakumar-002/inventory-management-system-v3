from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
import os
import io
from openpyxl import Workbook

# Initialize Flask app
app = Flask(__name__)

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production-2025-xyz')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///ims_complete.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ECHO'] = os.environ.get('FLASK_ENV') == 'development'
app.config['WTF_CSRF_ENABLED'] = False  # Disable CSRF for development

# Session Configuration
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db)

# ==================== TEMPLATE FILTERS & CONTEXT PROCESSORS ====================

@app.template_filter('currency')
def format_currency(value):
    """Format value as Indian Rupee currency"""
    if value is None:
        value = 0
    try:
        return f"₹{float(value):,.2f}"
    except (ValueError, TypeError):
        return "₹0.00"

@app.template_filter('format_currency_indian')
def format_currency_indian(value):
    """Format value as Indian Rupee currency (alias)"""
    if value is None:
        value = 0
    try:
        return f"₹{float(value):,.2f}"
    except (ValueError, TypeError):
        return "₹0.00"

@app.template_filter('format_date')
def format_date(value):
    """Format datetime to readable date"""
    if value is None:
        return 'N/A'
    try:
        return value.strftime('%d %b %Y')
    except:
        return str(value)

@app.template_filter('format_datetime')
def format_datetime(value):
    """Format datetime to readable date and time"""
    if value is None:
        return 'N/A'
    try:
        return value.strftime('%d %b %Y, %I:%M %p')
    except:
        return str(value)

@app.template_filter('format_number')
def format_number(value):
    """Format number with thousand separators"""
    if value is None:
        value = 0
    try:
        return f"{float(value):,.0f}"
    except (ValueError, TypeError):
        return "0"

@app.template_filter('pluralize')
def pluralize(count, singular='', plural='s'):
    """Return plural suffix if count is not 1"""
    if count == 1:
        return singular
    return plural

# Register filters explicitly for Jinja2
app.jinja_env.filters['currency'] = format_currency
app.jinja_env.filters['format_currency_indian'] = format_currency_indian
app.jinja_env.filters['format_date'] = format_date
app.jinja_env.filters['format_datetime'] = format_datetime
app.jinja_env.filters['format_number'] = format_number
app.jinja_env.filters['pluralize'] = pluralize

# Add context processor for global template variables
@app.context_processor
def utility_processor():
    """Add utility functions to all templates"""
    def format_price(amount):
        if amount is None:
            amount = 0
        return f"₹{float(amount):,.2f}"
    
    def format_qty(quantity):
        if quantity is None:
            return 0
        return int(quantity)
    
    def format_currency_indian_func(value):
        """Format value as Indian Rupee currency - function version"""
        if value is None:
            value = 0
        try:
            return f"₹{float(value):,.2f}"
        except (ValueError, TypeError):
            return "₹0.00"
    
    def csrf_token():
        return ''
    
    return dict(
        format_price=format_price,
        format_qty=format_qty,
        format_currency_indian=format_currency_indian_func,
        csrf_token=csrf_token,
        now=datetime.utcnow
    )
# ==================== DATABASE MODELS ====================

class User(db.Model):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(120), nullable=False)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), nullable=True)
    role = db.Column(db.String(50), default='user')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<User {self.username}>'


class Category(db.Model):
    __tablename__ = 'category'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True, index=True)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    products = db.relationship('Product', backref='category', lazy=True, cascade='all, delete-orphan')

    def __repr__(self):
        return f'<Category {self.name}>'


class Product(db.Model):
    __tablename__ = 'product'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, index=True)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    description = db.Column(db.Text)
    unit_price = db.Column(db.Float, nullable=False)
    quantity = db.Column(db.Integer, default=0)
    reorder_level = db.Column(db.Integer, default=10)
    sku = db.Column(db.String(100), unique=True, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def __repr__(self):
        return f'<Product {self.name}>'


class Supplier(db.Model):
    __tablename__ = 'supplier'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, index=True)
    contact_person = db.Column(db.String(100))
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    address = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Supplier {self.name}>'


class Customer(db.Model):
    __tablename__ = 'customer'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, index=True)
    email = db.Column(db.String(120))
    phone = db.Column(db.String(20))
    address = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Customer {self.name}>'


class Sale(db.Model):
    __tablename__ = 'sale'
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'))
    total_amount = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(50))
    status = db.Column(db.String(50), default='completed')
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    sale_items = db.relationship('SaleItem', backref='sale', lazy=True, cascade='all, delete-orphan')
    customer = db.relationship('Customer', backref='sales')

    def __repr__(self):
        return f'<Sale #{self.id}>'


class SaleItem(db.Model):
    __tablename__ = 'sale_item'
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey('sale.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    product = db.relationship('Product')


class PurchaseOrder(db.Model):
    __tablename__ = 'purchase_order'
    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'))
    total_amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(50), default='pending', index=True)
    order_date = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    expected_delivery = db.Column(db.DateTime)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    purchase_items = db.relationship('PurchaseItem', backref='purchase_order', lazy=True, cascade='all, delete-orphan')
    supplier = db.relationship('Supplier', backref='purchase_orders')

    def __repr__(self):
        return f'<PurchaseOrder #{self.id}>'


class PurchaseItem(db.Model):
    __tablename__ = 'purchase_item'
    id = db.Column(db.Integer, primary_key=True)
    purchase_order_id = db.Column(db.Integer, db.ForeignKey('purchase_order.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    subtotal = db.Column(db.Float, nullable=False)
    product = db.relationship('Product')


class StockHistory(db.Model):
    __tablename__ = 'stock_history'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False, index=True)
    change_type = db.Column(db.String(50), nullable=False)
    quantity_change = db.Column(db.Integer, nullable=False)
    previous_quantity = db.Column(db.Integer)
    new_quantity = db.Column(db.Integer)
    reference_id = db.Column(db.Integer)
    reference_type = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    product = db.relationship('Product')

    def __repr__(self):
        return f'<StockHistory {self.product_id} - {self.change_type}>'
# ==================== AUTHENTICATION DECORATOR ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== HOME & AUTHENTICATION ROUTES ====================

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/home')
def home_page():
    """Alternative home route"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    # If already logged in, redirect to dashboard
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Validate input
        if not username or not password:
            flash('Please enter both username and password', 'warning')
            return render_template('login.html')
        
        # Find user
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            # Set session data
            session.permanent = True
            session['user_id'] = user.id
            session['username'] = user.username
            session['name'] = user.name
            session['role'] = user.role
            
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'danger')
            return render_template('login.html')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    try:
        total_products = Product.query.count()
        total_categories = Category.query.count()
        total_suppliers = Supplier.query.count()
        total_customers = Customer.query.count()
        
        low_stock_products = Product.query.filter(Product.quantity <= Product.reorder_level).all()
        recent_sales = Sale.query.order_by(Sale.created_at.desc()).limit(5).all()
        
        today = datetime.utcnow().date()
        today_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
            db.func.date(Sale.created_at) == today
        ).scalar() or 0
        
        first_day_month = today.replace(day=1)
        month_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
            Sale.created_at >= first_day_month
        ).scalar() or 0
        
        total_value = db.session.query(
            db.func.sum(Product.quantity * Product.unit_price)
        ).scalar() or 0
        
        return render_template('inventory.html', 
                             total_products=total_products,
                             total_categories=total_categories,
                             total_suppliers=total_suppliers,
                             total_customers=total_customers,
                             low_stock_products=low_stock_products,
                             recent_sales=recent_sales,
                             today_sales=today_sales,
                             month_sales=month_sales,
                             total_value=total_value)
    except Exception as e:
        app.logger.error(f"Dashboard error: {str(e)}")
        db.session.rollback()
        return render_template('inventory.html',
                             total_products=0,
                             total_categories=0,
                             total_suppliers=0,
                             total_customers=0,
                             low_stock_products=[],
                             recent_sales=[],
                             today_sales=0,
                             month_sales=0,
                             total_value=0)
# ==================== CATEGORY ROUTES ====================

@app.route('/categories')
@login_required
def categories():
    all_categories = Category.query.order_by(Category.name).all()
    return render_template('manage_category.html', categories=all_categories)


@app.route('/manage_category')
@login_required
def manage_category():
    return redirect(url_for('categories'))


@app.route('/category/add', methods=['POST'])
@login_required
def add_category():
    name = request.form.get('name')
    description = request.form.get('description')
    
    if Category.query.filter_by(name=name).first():
        flash('Category already exists!', 'warning')
        return redirect(url_for('categories'))
    
    new_category = Category(name=name, description=description)
    db.session.add(new_category)
    db.session.commit()
    
    flash('Category added successfully!', 'success')
    return redirect(url_for('categories'))


@app.route('/category/edit/<int:id>', methods=['POST'])
@login_required
def edit_category(id):
    category = Category.query.get_or_404(id)
    category.name = request.form.get('name')
    category.description = request.form.get('description')
    
    db.session.commit()
    flash('Category updated successfully!', 'success')
    return redirect(url_for('categories'))


@app.route('/category/delete/<int:id>')
@login_required
def delete_category(id):
    category = Category.query.get_or_404(id)
    
    if category.products:
        flash('Cannot delete category with existing products!', 'danger')
        return redirect(url_for('categories'))
    
    db.session.delete(category)
    db.session.commit()
    flash('Category deleted successfully!', 'success')
    return redirect(url_for('categories'))


# ==================== PRODUCT ROUTES ====================

@app.route('/products')
@login_required
def products():
    all_products = Product.query.join(Category).order_by(Product.name).all()
    all_categories = Category.query.order_by(Category.name).all()
    return render_template('manage_product.html', products=all_products, categories=all_categories)


@app.route('/manage_product')
@login_required
def manage_product():
    return redirect(url_for('products'))


@app.route('/product_mgt')
@login_required
def product_mgt():
    return redirect(url_for('products'))


@app.route('/product/add', methods=['POST'])
@login_required
def add_product():
    name = request.form.get('name')
    category_id = request.form.get('category_id')
    description = request.form.get('description')
    unit_price = float(request.form.get('unit_price'))
    quantity = int(request.form.get('quantity', 0))
    reorder_level = int(request.form.get('reorder_level', 10))
    sku = request.form.get('sku')
    
    new_product = Product(
        name=name,
        category_id=category_id,
        description=description,
        unit_price=unit_price,
        quantity=quantity,
        reorder_level=reorder_level,
        sku=sku
    )
    
    db.session.add(new_product)
    db.session.commit()
    
    if quantity > 0:
        history = StockHistory(
            product_id=new_product.id,
            change_type='initial_stock',
            quantity_change=quantity,
            previous_quantity=0,
            new_quantity=quantity,
            created_by=session.get('user_id')
        )
        db.session.add(history)
        db.session.commit()
    
    flash('Product added successfully!', 'success')
    return redirect(url_for('products'))


@app.route('/product/edit/<int:id>', methods=['POST'])
@login_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    
    product.name = request.form.get('name')
    product.category_id = request.form.get('category_id')
    product.description = request.form.get('description')
    product.unit_price = float(request.form.get('unit_price'))
    product.reorder_level = int(request.form.get('reorder_level'))
    product.sku = request.form.get('sku')
    product.updated_at = datetime.utcnow()
    
    db.session.commit()
    flash('Product updated successfully!', 'success')
    return redirect(url_for('products'))


@app.route('/product/delete/<int:id>')
@login_required
def delete_product(id):
    product = Product.query.get_or_404(id)
    db.session.delete(product)
    db.session.commit()
    flash('Product deleted successfully!', 'success')
    return redirect(url_for('products'))


@app.route('/api/product/<int:id>')
@login_required
def get_product(id):
    product = Product.query.get_or_404(id)
    return jsonify({
        'id': product.id,
        'name': product.name,
        'unit_price': product.unit_price,
        'quantity': product.quantity,
        'category': product.category.name,
        'sku': product.sku
    })


# ==================== STOCK MANAGEMENT ROUTES ====================

@app.route('/stock')
@login_required
def stock_management():
    all_products = Product.query.join(Category).order_by(Product.name).all()
    return render_template('manage_stock.html', products=all_products)


@app.route('/manage_stock')
@login_required
def manage_stock():
    return redirect(url_for('stock_management'))


@app.route('/stock/adjust/<int:id>', methods=['POST'])
@login_required
def adjust_stock(id):
    product = Product.query.get_or_404(id)
    
    adjustment_type = request.form.get('adjustment_type')
    quantity = int(request.form.get('quantity'))
    reason = request.form.get('reason', 'manual_adjustment')
    
    previous_qty = product.quantity
    
    if adjustment_type == 'add':
        product.quantity += quantity
        quantity_change = quantity
    else:
        product.quantity -= quantity
        quantity_change = -quantity
    
    product.updated_at = datetime.utcnow()
    
    history = StockHistory(
        product_id=product.id,
        change_type=reason,
        quantity_change=quantity_change,
        previous_quantity=previous_qty,
        new_quantity=product.quantity,
        created_by=session.get('user_id')
    )
    
    db.session.add(history)
    db.session.commit()
    
    flash('Stock adjusted successfully!', 'success')
    return redirect(url_for('stock_management'))


@app.route('/stock/history')
@login_required
def stock_history():
    history = StockHistory.query.join(Product).order_by(StockHistory.created_at.desc()).limit(100).all()
    return render_template('inventory-history.html', history=history)


@app.route('/inventory-history')
@login_required
def inventory_history():
    return redirect(url_for('stock_history'))
# ==================== SUPPLIER ROUTES ====================

@app.route('/suppliers')
@login_required
def suppliers():
    all_suppliers = Supplier.query.order_by(Supplier.name).all()
    return render_template('suppliers.html', suppliers=all_suppliers)


@app.route('/manage_supplier')
@login_required
def manage_supplier():
    return redirect(url_for('suppliers'))


@app.route('/supplier/add', methods=['POST'])
@login_required
def add_supplier():
    name = request.form.get('name')
    contact_person = request.form.get('contact_person')
    email = request.form.get('email')
    phone = request.form.get('phone')
    address = request.form.get('address')
    
    new_supplier = Supplier(
        name=name,
        contact_person=contact_person,
        email=email,
        phone=phone,
        address=address
    )
    
    db.session.add(new_supplier)
    db.session.commit()
    
    flash('Supplier added successfully!', 'success')
    return redirect(url_for('suppliers'))


@app.route('/supplier/edit/<int:id>', methods=['POST'])
@login_required
def edit_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    
    supplier.name = request.form.get('name')
    supplier.contact_person = request.form.get('contact_person')
    supplier.email = request.form.get('email')
    supplier.phone = request.form.get('phone')
    supplier.address = request.form.get('address')
    
    db.session.commit()
    flash('Supplier updated successfully!', 'success')
    return redirect(url_for('suppliers'))


@app.route('/supplier/delete/<int:id>')
@login_required
def delete_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    db.session.delete(supplier)
    db.session.commit()
    flash('Supplier deleted successfully!', 'success')
    return redirect(url_for('suppliers'))


# ==================== CUSTOMER ROUTES ====================

@app.route('/customers')
@login_required
def customers():
    all_customers = Customer.query.order_by(Customer.name).all()
    return render_template('customers.html', customers=all_customers)


@app.route('/customer/add', methods=['POST'])
@login_required
def add_customer_form():
    name = request.form.get('name')
    email = request.form.get('email')
    phone = request.form.get('phone')
    address = request.form.get('address')
    
    new_customer = Customer(name=name, email=email, phone=phone, address=address)
    db.session.add(new_customer)
    db.session.commit()
    
    flash('Customer added successfully!', 'success')
    return redirect(url_for('customers'))


@app.route('/customer/edit/<int:id>', methods=['POST'])
@login_required
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    
    customer.name = request.form.get('name')
    customer.email = request.form.get('email')
    customer.phone = request.form.get('phone')
    customer.address = request.form.get('address')
    
    db.session.commit()
    flash('Customer updated successfully!', 'success')
    return redirect(url_for('customers'))


@app.route('/customer/delete/<int:id>')
@login_required
def delete_customer(id):
    customer = Customer.query.get_or_404(id)
    db.session.delete(customer)
    db.session.commit()
    flash('Customer deleted successfully!', 'success')
    return redirect(url_for('customers'))


@app.route('/api/customers')
@login_required
def get_customers():
    customers = Customer.query.all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'email': c.email,
        'phone': c.phone
    } for c in customers])


@app.route('/api/customer/add', methods=['POST'])
@login_required
def add_customer():
    data = request.get_json()
    
    new_customer = Customer(
        name=data.get('name'),
        email=data.get('email'),
        phone=data.get('phone'),
        address=data.get('address')
    )
    
    db.session.add(new_customer)
    db.session.commit()
    
    return jsonify({'success': True, 'customer_id': new_customer.id})


# ==================== SALES ROUTES ====================

@app.route('/sales')
@login_required
def sales():
    all_sales = Sale.query.order_by(Sale.created_at.desc()).all()
    return render_template('sales.html', sales=all_sales)


@app.route('/sale/new')
@login_required
def new_sale():
    products = Product.query.filter(Product.quantity > 0).all()
    customers = Customer.query.all()
    return render_template('new_sale.html', products=products, customers=customers)


@app.route('/new_sale')
@login_required
def new_sale_alt():
    return redirect(url_for('new_sale'))


@app.route('/sale/create', methods=['POST'])
@login_required
def create_sale():
    customer_id = request.form.get('customer_id') or None
    payment_method = request.form.get('payment_method')
    
    product_ids = request.form.getlist('product_id[]')
    quantities = request.form.getlist('quantity[]')
    unit_prices = request.form.getlist('unit_price[]')
    
    if not product_ids:
        flash('Please add at least one product to the sale!', 'warning')
        return redirect(url_for('new_sale'))
    
    total_amount = 0
    sale_items = []
    
    for i in range(len(product_ids)):
        product_id = int(product_ids[i])
        quantity = int(quantities[i])
        unit_price = float(unit_prices[i])
        subtotal = quantity * unit_price
        
        product = Product.query.get(product_id)
        
        if product.quantity < quantity:
            flash(f'Insufficient stock for {product.name}!', 'danger')
            return redirect(url_for('new_sale'))
        
        sale_items.append({
            'product_id': product_id,
            'quantity': quantity,
            'unit_price': unit_price,
            'subtotal': subtotal
        })
        
        total_amount += subtotal
    
    new_sale = Sale(
        customer_id=customer_id,
        total_amount=total_amount,
        payment_method=payment_method,
        created_by=session.get('user_id')
    )
    
    db.session.add(new_sale)
    db.session.flush()
    
    for item in sale_items:
        sale_item = SaleItem(
            sale_id=new_sale.id,
            product_id=item['product_id'],
            quantity=item['quantity'],
            unit_price=item['unit_price'],
            subtotal=item['subtotal']
        )
        db.session.add(sale_item)
        
        product = Product.query.get(item['product_id'])
        previous_qty = product.quantity
        product.quantity -= item['quantity']
        product.updated_at = datetime.utcnow()
        
        history = StockHistory(
            product_id=product.id,
            change_type='sale',
            quantity_change=-item['quantity'],
            previous_quantity=previous_qty,
            new_quantity=product.quantity,
            reference_id=new_sale.id,
            reference_type='sale',
            created_by=session.get('user_id')
        )
        db.session.add(history)
    
    db.session.commit()
    
    flash('Sale created successfully!', 'success')
    return redirect(url_for('view_invoice', id=new_sale.id))


@app.route('/sale/invoice/<int:id>')
@login_required
def view_invoice(id):
    sale = Sale.query.get_or_404(id)
    return render_template('view_invoice.html', sale=sale)


@app.route('/invoice/<int:id>')
@login_required
def invoice(id):
    return redirect(url_for('view_invoice', id=id))


@app.route('/sale/delete/<int:id>')
@login_required
def delete_sale(id):
    sale = Sale.query.get_or_404(id)
    
    for item in sale.sale_items:
        product = Product.query.get(item.product_id)
        product.quantity += item.quantity
        product.updated_at = datetime.utcnow()
    
    db.session.delete(sale)
    db.session.commit()
    
    flash('Sale deleted and stock restored!', 'success')
    return redirect(url_for('sales'))
# ==================== PURCHASE ORDER ROUTES ====================

@app.route('/purchases')
@login_required
def purchase_orders():
    orders = PurchaseOrder.query.order_by(PurchaseOrder.order_date.desc()).all()
    return render_template('purchase_orders.html', orders=orders)


@app.route('/purchase_orders')
@login_required
def purchase_orders_alt():
    return redirect(url_for('purchase_orders'))


@app.route('/purchase/new')
@login_required
def new_purchase():
    products = Product.query.all()
    suppliers = Supplier.query.all()
    return render_template('new_purchase_order.html', products=products, suppliers=suppliers)


@app.route('/new_purchase_order')
@login_required
def new_purchase_order():
    return redirect(url_for('new_purchase'))


@app.route('/purchase/create', methods=['POST'])
@login_required
def create_purchase():
    supplier_id = request.form.get('supplier_id')
    expected_delivery = request.form.get('expected_delivery')
    
    product_ids = request.form.getlist('product_id[]')
    quantities = request.form.getlist('quantity[]')
    unit_prices = request.form.getlist('unit_price[]')
    
    if not product_ids:
        flash('Please add at least one product to the purchase order!', 'warning')
        return redirect(url_for('new_purchase'))
    
    total_amount = 0
    purchase_items = []
    
    for i in range(len(product_ids)):
        if not product_ids[i]:  # Skip empty rows
            continue
            
        product_id = int(product_ids[i])
        quantity = int(quantities[i])
        unit_price = float(unit_prices[i])
        subtotal = quantity * unit_price
        
        purchase_items.append({
            'product_id': product_id,
            'quantity': quantity,
            'unit_price': unit_price,
            'subtotal': subtotal
        })
        
        total_amount += subtotal
    
    new_purchase = PurchaseOrder(
        supplier_id=supplier_id,
        total_amount=total_amount,
        expected_delivery=datetime.strptime(expected_delivery, '%Y-%m-%d') if expected_delivery else None,
        created_by=session.get('user_id')
    )
    
    db.session.add(new_purchase)
    db.session.flush()
    
    for item in purchase_items:
        purchase_item = PurchaseItem(
            purchase_order_id=new_purchase.id,
            product_id=item['product_id'],
            quantity=item['quantity'],
            unit_price=item['unit_price'],
            subtotal=item['subtotal']
        )
        db.session.add(purchase_item)
    
    db.session.commit()
    
    flash('Purchase order created successfully!', 'success')
    return redirect(url_for('purchase_orders'))


@app.route('/purchase/receive/<int:id>')
@login_required
def receive_purchase(id):
    purchase = PurchaseOrder.query.get_or_404(id)
    
    if purchase.status == 'received':
        flash('Purchase order already received!', 'warning')
        return redirect(url_for('purchase_orders'))
    
    for item in purchase.purchase_items:
        product = Product.query.get(item.product_id)
        previous_qty = product.quantity
        product.quantity += item.quantity
        product.updated_at = datetime.utcnow()
        
        history = StockHistory(
            product_id=product.id,
            change_type='purchase',
            quantity_change=item.quantity,
            previous_quantity=previous_qty,
            new_quantity=product.quantity,
            reference_id=purchase.id,
            reference_type='purchase',
            created_by=session.get('user_id')
        )
        db.session.add(history)
    
    purchase.status = 'received'
    db.session.commit()
    
    flash('Purchase order received and stock updated!', 'success')
    return redirect(url_for('purchase_orders'))


@app.route('/purchase/delete/<int:id>')
@login_required
def delete_purchase(id):
    purchase = PurchaseOrder.query.get_or_404(id)
    
    if purchase.status == 'received':
        flash('Cannot delete received purchase order!', 'danger')
        return redirect(url_for('purchase_orders'))
    
    db.session.delete(purchase)
    db.session.commit()
    
    flash('Purchase order deleted successfully!', 'success')
    return redirect(url_for('purchase_orders'))


# ==================== REPORTS ROUTES ====================

@app.route('/reports')
@login_required
def reports():
    today = datetime.utcnow().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    today_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
        db.func.date(Sale.created_at) == today
    ).scalar() or 0
    
    week_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
        Sale.created_at >= week_ago
    ).scalar() or 0
    
    month_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(
        Sale.created_at >= month_ago
    ).scalar() or 0
    
    total_sales = db.session.query(db.func.sum(Sale.total_amount)).scalar() or 0
    
    low_stock = Product.query.filter(Product.quantity <= Product.reorder_level).all()
    
    top_products = db.session.query(
        Product.name,
        db.func.sum(SaleItem.quantity).label('total_sold'),
        db.func.sum(SaleItem.subtotal).label('total_revenue')
    ).join(SaleItem).group_by(Product.id).order_by(db.desc('total_sold')).limit(10).all()
    
    recent_activities = StockHistory.query.order_by(StockHistory.created_at.desc()).limit(10).all()
    
    return render_template('reports.html',
                         today_sales=today_sales,
                         week_sales=week_sales,
                         month_sales=month_sales,
                         total_sales=total_sales,
                         low_stock=low_stock,
                         top_products=top_products,
                         recent_activities=recent_activities)


@app.route('/reports/export/sales')
@login_required
def export_sales():
    sales = Sale.query.order_by(Sale.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    ws.append(['Sale ID', 'Date', 'Customer', 'Total Amount', 'Payment Method', 'Status', 'Created By'])
    
    for sale in sales:
        customer_name = 'Walk-in Customer'
        if sale.customer_id:
            customer = Customer.query.get(sale.customer_id)
            if customer:
                customer_name = customer.name
        
        creator = User.query.get(sale.created_by) if sale.created_by else None
        creator_name = creator.name if creator else 'Unknown'
        
        ws.append([
            sale.id,
            sale.created_at.strftime('%Y-%m-%d %H:%M'),
            customer_name,
            sale.total_amount,
            sale.payment_method,
            sale.status,
            creator_name
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/reports/export/products')
@login_required
def export_products():
    products = Product.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products Report"
    
    ws.append(['Product ID', 'Name', 'Category', 'SKU', 'Price', 'Quantity', 'Reorder Level', 'Value'])
    
    for product in products:
        ws.append([
            product.id,
            product.name,
            product.category.name,
            product.sku or 'N/A',
            product.unit_price,
            product.quantity,
            product.reorder_level,
            product.quantity * product.unit_price
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'products_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/reports/export/inventory')
@login_required
def export_inventory():
    history = StockHistory.query.join(Product).order_by(StockHistory.created_at.desc()).limit(500).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory History"
    
    ws.append(['Date', 'Product', 'Change Type', 'Quantity Change', 'Previous Qty', 'New Qty', 'Reference'])
    
    for h in history:
        ws.append([
            h.created_at.strftime('%Y-%m-%d %H:%M'),
            h.product.name,
            h.change_type,
            h.quantity_change,
            h.previous_quantity,
            h.new_quantity,
            f"{h.reference_type}#{h.reference_id}" if h.reference_type else 'Manual'
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventory_history_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )
# ==================== PROFILE ROUTES ====================

@app.route('/profile')
@login_required
def profile():
    user = User.query.get(session.get('user_id'))
    return render_template('manage_profile.html', user=user)


@app.route('/manage_profile')
@login_required
def manage_profile():
    return redirect(url_for('profile'))


@app.route('/profile/update', methods=['POST'])
@login_required
def update_profile():
    user = User.query.get(session.get('user_id'))
    
    user.name = request.form.get('name')
    user.email = request.form.get('email')
    
    db.session.commit()
    session['name'] = user.name
    
    flash('Profile updated successfully!', 'success')
    return redirect(url_for('profile'))


@app.route('/profile/change-password', methods=['POST'])
@login_required
def change_password():
    user = User.query.get(session.get('user_id'))
    
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    if not user.check_password(current_password):
        flash('Current password is incorrect!', 'danger')
        return redirect(url_for('profile'))
    
    if new_password != confirm_password:
        flash('New passwords do not match!', 'danger')
        return redirect(url_for('profile'))
    
    if len(new_password) < 6:
        flash('Password must be at least 6 characters long!', 'danger')
        return redirect(url_for('profile'))
    
    user.set_password(new_password)
    db.session.commit()
    
    flash('Password changed successfully!', 'success')
    return redirect(url_for('profile'))


@app.route('/update_password', methods=['GET', 'POST'])
@login_required
def update_password():
    if request.method == 'POST':
        return change_password()
    return render_template('update_password.html')


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found_error(error):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Not found'}), 404
    flash('Page not found!', 'warning')
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    app.logger.error(f'Internal error: {error}')
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error'}), 500
    flash('An internal error occurred. Please try again.', 'danger')
    if 'user_id' in session and request.path != '/dashboard':
        return redirect(url_for('dashboard'))
    return render_template('login.html'), 500


# ==================== CLI COMMANDS ====================

@app.cli.command()
def init_db():
    """Initialize the database with tables and default admin user"""
    db.create_all()
    
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            name='Administrator',
            email='admin@ims.com',
            role='admin'
        )
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print('✓ Database initialized with default admin user')
        print('  Username: admin')
        print('  Password: admin123')
    else:
        print('✓ Database already initialized')


@app.cli.command()
def seed_db():
    """Seed database with sample data"""
    categories = [
        Category(name='Electronics', description='Electronic items and gadgets'),
        Category(name='Furniture', description='Office and home furniture'),
        Category(name='Stationery', description='Office supplies and stationery'),
        Category(name='Food & Beverages', description='Food items and drinks')
    ]
    
    for cat in categories:
        if not Category.query.filter_by(name=cat.name).first():
            db.session.add(cat)
    
    db.session.commit()
    
    electronics = Category.query.filter_by(name='Electronics').first()
    if electronics:
        products = [
            Product(name='Laptop', category_id=electronics.id, unit_price=45000, quantity=10, sku='ELEC001', description='Business laptop'),
            Product(name='Mouse', category_id=electronics.id, unit_price=500, quantity=50, sku='ELEC002', description='Wireless mouse'),
            Product(name='Keyboard', category_id=electronics.id, unit_price=1500, quantity=30, sku='ELEC003', description='Mechanical keyboard')
        ]
        
        for prod in products:
            if not Product.query.filter_by(sku=prod.sku).first():
                db.session.add(prod)
    
    db.session.commit()
    print('✓ Database seeded with sample data')


# ==================== APPLICATION ENTRY POINT ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Create default admin if doesn't exist
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(username='admin', name='Administrator', email='admin@ims.com', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('Default admin user created: admin/admin123')
    
    debug_mode = os.environ.get('FLASK_ENV') == 'development'
    port = int(os.environ.get('PORT', 5000))
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
